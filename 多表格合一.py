import os
import pandas as pd
from tkinter import Tk, filedialog
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def categorize_column(header_name):
    """
    根据列名文本，将其分类到四大标准类别中
    使用正则表达式进行模糊匹配，兼容各种奇怪的命名
    """
    # 转为字符串并去除首尾空格
    if pd.isna(header_name):
        return None
    header_str = str(header_name).strip()

    # 定义每个标准类别对应的正则关键词模式
    # 只要匹配到任意一个关键词即判定为该类
    patterns = {
        "姓名": [r"姓名", r"名字", r"人员", r"干部", r"姓\s*名"], # 兼容 "姓 名"
        "身份证号": [r"身份证", r"证件号", r"身份证明"],
        "职务或单位": [r"单位", r"职务", r"岗位", r"职位", r"任职", r"部门", r"公司", r"机构", r"职级", r"现.?任.?职"],
        "备注": [r"备注", r"说明", r"注释", r"额外", r"补充", r"其他"]
    }

    for category, regex_list in patterns.items():
        for pattern in regex_list:
            # re.search 查找字符串中是否包含该模式
            if re.search(pattern, header_str, re.IGNORECASE):
                return category
    return None

def merge_excel_row_by_row():
    root = Tk()
    root.withdraw() 
    
    print("正在弹出文件夹选择对话框...")
    input_folder = filedialog.askdirectory(title="请选择包含Excel表格的文件夹")
    
    if not input_folder:
        print(" 未选择文件夹，程序退出。")
        return
    
    all_rows = []
    standard_columns = ["序号", "姓名", "身份证号", "职务或单位", "备注"]
    
    print(f"\n✅ 已选择文件夹: {input_folder}")
    print("开始处理文件...")

    excel_files = [f for f in os.listdir(input_folder) if f.lower().endswith(('.xlsx', '.xls', '.xlsm'))]
    
    if not excel_files:
        print("❌ 在指定文件夹中未找到Excel文件。")
        return

    for filename in excel_files:
        file_path = os.path.join(input_folder, filename)
        try:
            print(f"\n📄 正在处理: {filename}")
            
            df_raw = pd.read_excel(file_path, skiprows=1, header=0, dtype=str)
            df_raw.dropna(how='all', inplace=True)
            
            if df_raw.empty:
                print(f"   ⚠️ 警告: 文件为空，已跳过。")
                continue
            
            col_mapping = {}
            for col in df_raw.columns:
                std_col = categorize_column(col)
                if std_col:
                    col_mapping[col] = std_col
                    print(f"   🔗 映射: '{col}' -> '{std_col}'")
            
            if not col_mapping:
                print(f"   ❌ 错误: 未识别出有效列，请检查表头格式。")
                continue
            
            # --- 添加独立序号 ---
            row_count = len(df_raw)
            serial_numbers = list(range(1, row_count + 1))
            
            # --- 逐行处理 ---
            for idx, row in df_raw.iterrows():
                new_row = {col: "" for col in standard_columns}
                new_row["序号"] = serial_numbers[idx] # 填入序号
                
                for raw_col, std_col in col_mapping.items():
                    if raw_col in row and pd.notna(row[raw_col]):
                        cell_value = str(row[raw_col]).strip()
                        # 防止重复填充覆盖（例如多个列都被识别为备注）
                        if not new_row[std_col]: 
                            new_row[std_col] = cell_value
                
                # 清理备注字段，去掉可能的'nan'
                if "nan" in new_row["备注"]:
                    new_row["备注"] = ""
                new_row["备注"] += f" | 来源: {filename}"
                
                all_rows.append(new_row)
            
            print(f"   ✅ 成功处理 {row_count} 行数据。")

        except Exception as e:
            print(f"❌ 处理失败 {filename}: {str(e)}")

    # --- 保存与样式设置 ---
    if all_rows:
        try:
            output_df = pd.DataFrame(all_rows, columns=standard_columns)
            
            output_path = os.path.join(os.getcwd(), "【最终版】联审人员汇总结果.xlsx")
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                output_df.to_excel(writer, index=False, sheet_name='联审名单')
                
                workbook = writer.book
                worksheet = writer.sheets['联审名单']
                
                # 1. 设置全局字体为 14号
                font_style = Font(size=14)
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.font = font_style
                
                # 2. 自动调整列宽
                for i, col in enumerate(output_df.columns, 1):
                    max_length = max(
                        output_df[col].astype(str).map(len).max(), # 数据最长长度
                        len(str(col)) # 列名长度
                    )
                    adjusted_width = min(max_length + 2, 50) # 加2个字符边距，限制最大宽度
                    worksheet.column_dimensions[get_column_letter(i)].width = adjusted_width

            print(f"\n🎉 合并完成！共处理 {len(output_df)} 行数据。")
            print(f"📁 保存路径: {output_path}")
        except Exception as save_err:
            print(f"❌ 保存文件时出错: {save_err}")
    else:
        print("\n❌ 未能提取到任何数据，请检查文件格式是否正确。")

# --- 程序入口 ---
if __name__ == "__main__":
    try:
        import re # 导入正则模块
        merge_excel_row_by_row()
    except Exception as main_err:
        print(f"🚨 程序启动失败: {main_err}")
        import traceback
        traceback.print_exc()
    
    print("----------------------------------------")
    input("按回车键退出...") 
